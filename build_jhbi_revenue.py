from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

wb = Workbook()
wb.remove(wb.active)

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY     = "06185F"
COBALT   = "085CE5"
COBALT_D = "073FA8"
WHITE    = "FFFFFF"
LT_BLUE  = "DCE6F1"
VLT_BLUE = "EBF3FB"
LT_GRAY  = "F5F5F5"
MID_GRAY = "D9D9D9"
YELLOW   = "FFFF00"
IN_BLUE  = "0000FF"
FX_BLACK = "000000"
XS_GREEN = "008000"
GREEN_BG = "E2EFDA"
AMBER    = "FFC000"
RED_BG   = "FFE0E0"

# ── Style helpers ─────────────────────────────────────────────────────────────
def fnt(bold=False, color=FX_BLACK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fill(c): return PatternFill("solid", start_color=c, end_color=c)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def rgt(): return Alignment(horizontal="right",  vertical="center")

th  = Side(style="thin",   color="BFBFBF")
med = Side(style="medium", color="595959")

def bdr(left=True, right=True, top=True, bottom=True):
    return Border(
        left=th if left else None,
        right=th if right else None,
        top=th if top else None,
        bottom=th if bottom else None,
    )

def bot_bdr(): return Border(bottom=med)

def hdr_row(ws, row, ncols, label, bg=NAVY, size=10):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = fnt(bold=True, color=WHITE, size=size)
        cell.alignment = ctr()
        cell.border = bdr()
    ws.cell(row=row, column=1).value = label
    ws.cell(row=row, column=1).alignment = lft()

FMT_USD   = '$#,##0;($#,##0);"-"'
FMT_USD_K = '$#,##0,"K";($#,##0,"K");"-"'
FMT_PCT   = '0.0%;(0.0%);"-"'
FMT_INT   = '#,##0;(#,##0);"-"'

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — ASSUMPTIONS
# ─────────────────────────────────────────────────────────────────────────────
ws_a = wb.create_sheet("Assumptions")
ws_a.sheet_view.showGridLines = False
ws_a.freeze_panes = "B4"

col_w = {"A":36,"B":16,"C":16,"D":16,"E":16,"F":14,"G":28}
for col, w in col_w.items(): ws_a.column_dimensions[col].width = w

# Title
ws_a.row_dimensions[1].height = 30
ws_a.merge_cells("A1:G1")
c = ws_a["A1"]
c.value = "JHBI Revenue Model  —  Assumptions & Inputs"
c.font  = fnt(bold=True, color=WHITE, size=14)
c.fill  = fill(NAVY)
c.alignment = lft()

# Legend
ws_a.row_dimensions[2].height = 16
ws_a.merge_cells("A2:G2")
c = ws_a["A2"]
c.value = "   ■ Blue = Hardcoded input (editable)    ■ Black = Formula    ■ Yellow bg = Key assumption"
c.font  = fnt(color="595959", size=9)
c.fill  = fill(LT_GRAY)
c.alignment = lft()

r = 4  # start row

# ── Section: Market Parameters ───────────────────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 1 — MARKET PARAMETERS", bg=COBALT); r+=1
ws_a.row_dimensions[r-1].height = 18

params = [
    ("Total Addressable FIs (JH Core)", 1660,  FMT_INT,  True,  "JH ~1,660 contracted core banking FIs (banks + CUs)"),
    ("CU Share of Base",                0.44,  FMT_PCT,  True,  "Est. Symitar/CU share; internal estimate"),
    ("Bank Share of Base",              None,  FMT_PCT,  False, "Complement of CU share (formula-driven)"),
    ("CU Discount Rate",                0.12,  FMT_PCT,  True,  "Pricing accommodation for cooperative/NCUA budget cycles"),
    ("Zelle Penetration (% of FIs)",    0.157, FMT_PCT,  True,  "260 of ~1,660 JH core FIs currently Zelle-enabled via JHA PayCenter"),
    ("Annual Subscription Churn Rate",  0.08,  FMT_PCT,  True,  "Annual FI churn on DS app subscriptions; adjust as needed"),
]
for label, val, fmt, is_input, note in params:
    ws_a.row_dimensions[r].height = 18
    ws_a.cell(r,1).value = label
    ws_a.cell(r,1).font  = fnt()
    ws_a.cell(r,1).border = bdr()
    ws_a.cell(r,1).alignment = lft()

    if val is None:
        ws_a.cell(r,2).value  = "=1-B6"
        ws_a.cell(r,2).font   = fnt(color=FX_BLACK)
    else:
        ws_a.cell(r,2).value  = val
        ws_a.cell(r,2).font   = fnt(color=IN_BLUE if is_input else FX_BLACK, bold=is_input)
    ws_a.cell(r,2).number_format = fmt
    ws_a.cell(r,2).alignment = ctr()
    ws_a.cell(r,2).border    = bdr()
    if is_input: ws_a.cell(r,2).fill = fill(YELLOW)

    ws_a.cell(r,7).value     = note
    ws_a.cell(r,7).font      = fnt(color="595959", size=9, italic=True)
    ws_a.cell(r,7).alignment = lft()
    r += 1

r += 1  # blank

# ── Section: FI Counts by Tier ───────────────────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 2 — FI BASE BY ASSET TIER", bg=COBALT); r+=1
ws_a.row_dimensions[r-1].height = 18

# Column headers
for ci, h in enumerate(["Asset Tier","Total FIs","CU","Bank","Zelle-Enabled","","Notes"], 1):
    ws_a.cell(r,ci).value = h
    ws_a.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_a.cell(r,ci).fill  = fill(COBALT_D)
    ws_a.cell(r,ci).alignment = ctr()
    ws_a.cell(r,ci).border = bdr()
ws_a.row_dimensions[r].height = 18; r+=1

# Data rows — Total FIs are inputs; CU/Bank/Zelle are formulas
tier_data = [
    ("<$250M",    500),
    ("$250M–$1B", 800),
    ("$1B–$5B",   280),
    ("$5B+",       80),
]
tier_start = r
for i,(label,total) in enumerate(tier_data):
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_a.row_dimensions[r].height = 18
    ws_a.cell(r,1).value = label;  ws_a.cell(r,1).font = fnt(); ws_a.cell(r,1).fill = fill(bg); ws_a.cell(r,1).border = bdr(); ws_a.cell(r,1).alignment = lft()
    ws_a.cell(r,2).value = total;  ws_a.cell(r,2).font = fnt(color=IN_BLUE, bold=True); ws_a.cell(r,2).number_format = FMT_INT; ws_a.cell(r,2).fill = fill(YELLOW); ws_a.cell(r,2).border = bdr(); ws_a.cell(r,2).alignment = ctr()
    ws_a.cell(r,3).value = f"=ROUND(B{r}*$B$6,0)"; ws_a.cell(r,3).font = fnt(); ws_a.cell(r,3).number_format = FMT_INT; ws_a.cell(r,3).fill = fill(bg); ws_a.cell(r,3).border = bdr(); ws_a.cell(r,3).alignment = ctr()
    ws_a.cell(r,4).value = f"=B{r}-C{r}";            ws_a.cell(r,4).font = fnt(); ws_a.cell(r,4).number_format = FMT_INT; ws_a.cell(r,4).fill = fill(bg); ws_a.cell(r,4).border = bdr(); ws_a.cell(r,4).alignment = ctr()
    ws_a.cell(r,5).value = f"=ROUND(B{r}*$B$9,0)";  ws_a.cell(r,5).font = fnt(); ws_a.cell(r,5).number_format = FMT_INT; ws_a.cell(r,5).fill = fill(bg); ws_a.cell(r,5).border = bdr(); ws_a.cell(r,5).alignment = ctr()
    r += 1

# Totals
ws_a.row_dimensions[r].height = 18
for ci, formula in [(1,"TOTAL"),(2,f"=SUM(B{tier_start}:B{r-1})"),(3,f"=SUM(C{tier_start}:C{r-1})"),(4,f"=SUM(D{tier_start}:D{r-1})"),(5,f"=SUM(E{tier_start}:E{r-1})")]:
    c = ws_a.cell(r, ci)
    c.value = formula
    c.font  = fnt(bold=True, color=WHITE)
    c.fill  = fill(NAVY)
    c.border = bdr()
    c.alignment = ctr() if ci > 1 else lft()
    if ci > 1: c.number_format = FMT_INT
r += 2

# ── Section: App Pricing ─────────────────────────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 3 — DS APP PRICING  (Annual per FI, add-on to JHBI platform)", bg=COBALT); r+=1
ws_a.row_dimensions[r-1].height = 18

for ci, h in enumerate(["App","<$250M (Bank)","$250M–$1B (Bank)","$1B–$5B (Bank)","$5B+ (Bank)","","Notes"], 1):
    ws_a.cell(r,ci).value = h
    ws_a.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_a.cell(r,ci).fill  = fill(COBALT_D)
    ws_a.cell(r,ci).alignment = ctr()
    ws_a.cell(r,ci).border = bdr()
ws_a.row_dimensions[r].height = 18; r+=1

pricing_start = r
app_pricing = [
    ("Zelle Memo Intelligence",  5000,  8000, 15000, 25000, "Zelle-enabled FIs only (260 currently; ~15.7% of base)"),
    ("Churn Sentinel",           8000, 12000, 22000, 35000, "Universal — all FI types, all tiers"),
    ("CommercialSignal",         6000, 10000, 18000, 28000, "Universal — runs on ACH/payments data"),
    ("Gen. Wealth Deflection",   8000, 12000, 20000, 30000, "Universal — requires core + household data (Phase 3)"),
    ("Anomaly Detection",        8000, 15000, 25000, 40000, "Universal — real-time streaming (Phase 3)"),
]
for i,(app, t1, t2, t3, t4, note) in enumerate(app_pricing):
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_a.row_dimensions[r].height = 18
    for ci, v in [(1,app),(2,t1),(3,t2),(4,t3),(5,t4),(7,note)]:
        c = ws_a.cell(r,ci)
        c.value = v
        c.fill  = fill(bg) if ci != 7 else fill(WHITE)
        c.border = bdr() if ci != 7 else None
        c.alignment = lft() if ci in (1,7) else ctr()
        if ci == 1: c.font = fnt()
        elif ci == 7: c.font = fnt(color="595959", size=9, italic=True)
        else:
            c.font = fnt(color=IN_BLUE, bold=True)
            c.number_format = FMT_USD
            c.fill = fill(YELLOW)
    r += 1

r += 1

# ── Section: CU Effective Pricing ────────────────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 4 — CU EFFECTIVE PRICING  (Bank rate × (1 − CU Discount))", bg=COBALT_D); r+=1
ws_a.row_dimensions[r-1].height = 18

for ci, h in enumerate(["App","<$250M (CU)","$250M–$1B (CU)","$1B–$5B (CU)","$5B+ (CU)","",""], 1):
    ws_a.cell(r,ci).value = h
    ws_a.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_a.cell(r,ci).fill  = fill(COBALT_D)
    ws_a.cell(r,ci).alignment = ctr()
    ws_a.cell(r,ci).border = bdr()
ws_a.row_dimensions[r].height = 18; r+=1

cu_start = r
for i in range(5):
    pr = pricing_start + i
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_a.row_dimensions[r].height = 18
    ws_a.cell(r,1).value = f"=Assumptions!A{pr}"; ws_a.cell(r,1).font=fnt(color=XS_GREEN); ws_a.cell(r,1).fill=fill(bg); ws_a.cell(r,1).border=bdr(); ws_a.cell(r,1).alignment=lft()
    for ci, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
        c = ws_a.cell(r, ci)
        c.value  = f"=Assumptions!{col}{pr}*(1-$B$7)"
        c.font   = fnt()
        c.fill   = fill(bg)
        c.border = bdr()
        c.alignment = ctr()
        c.number_format = FMT_USD
    r += 1

r += 1

# ── Section: Blended Average Price per App ────────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 5 — BLENDED AVG SUBSCRIPTION PRICE  (Weighted by adoption tier mix, CU/Bank split)", bg=COBALT); r+=1
ws_a.row_dimensions[r-1].height = 18
ws_a.merge_cells(f"A{r}:G{r}")
ws_a.cell(r,1).value = "   Used in Revenue Model sheet. Weighted average accounts for tier mix of signing FIs and CU discount applied proportionally."
ws_a.cell(r,1).font  = fnt(color="595959", size=9, italic=True)
ws_a.cell(r,1).fill  = fill(LT_GRAY)
ws_a.cell(r,1).alignment = lft()
r += 1

for ci, h in enumerate(["App","Blended Avg Price / FI","Annual FI Target","Year 1 ARR Potential","","","Source / Notes"], 1):
    ws_a.cell(r,ci).value = h
    ws_a.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_a.cell(r,ci).fill  = fill(COBALT_D)
    ws_a.cell(r,ci).alignment = ctr()
    ws_a.cell(r,ci).border = bdr()
ws_a.row_dimensions[r].height = 18; r+=1

blend_start = r
# Blended prices pre-calculated (formula-derivable but hardcoded for clarity, flagged blue)
blended = [
    ("Zelle Memo Intelligence",  9700,  40, "Tier mix: 20%/<$250M, 50%/$250M-1B, 22.5%/$1B-5B, 7.5%/$5B+; CU share blended"),
    ("Churn Sentinel",          13700,  50, "Broader tier reach; higher avg due to $1B+ FI interest"),
    ("CommercialSignal",        11100,  35, "Payments data only; moderate tier skew"),
    ("Gen. Wealth Deflection",  13600,  24, "Core data dependency limits early adoption to mid/large tiers"),
    ("Anomaly Detection",       16800,  24, "Highest price; complex streaming model; targets Tier 3-4"),
]
for i,(app, price, target, note) in enumerate(blended):
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_a.row_dimensions[r].height = 18
    ws_a.cell(r,1).value = app;    ws_a.cell(r,1).font=fnt(); ws_a.cell(r,1).fill=fill(bg); ws_a.cell(r,1).border=bdr(); ws_a.cell(r,1).alignment=lft()
    ws_a.cell(r,2).value = price;  ws_a.cell(r,2).font=fnt(color=IN_BLUE,bold=True); ws_a.cell(r,2).fill=fill(YELLOW); ws_a.cell(r,2).number_format=FMT_USD; ws_a.cell(r,2).border=bdr(); ws_a.cell(r,2).alignment=ctr()
    ws_a.cell(r,3).value = target; ws_a.cell(r,3).font=fnt(color=IN_BLUE,bold=True); ws_a.cell(r,3).fill=fill(YELLOW); ws_a.cell(r,3).number_format=FMT_INT; ws_a.cell(r,3).border=bdr(); ws_a.cell(r,3).alignment=ctr()
    ws_a.cell(r,4).value = f"=B{r}*C{r}"; ws_a.cell(r,4).font=fnt(bold=True); ws_a.cell(r,4).fill=fill(GREEN_BG); ws_a.cell(r,4).number_format=FMT_USD; ws_a.cell(r,4).border=bdr(); ws_a.cell(r,4).alignment=ctr()
    ws_a.cell(r,7).value = note;   ws_a.cell(r,7).font=fnt(color="595959",size=9,italic=True); ws_a.cell(r,7).alignment=lft()
    r += 1

# Total row
ws_a.row_dimensions[r].height = 18
ws_a.cell(r,1).value = "TOTAL — Year 1 ARR Potential (all apps launched)"; ws_a.cell(r,1).font=fnt(bold=True,color=WHITE); ws_a.cell(r,1).fill=fill(NAVY); ws_a.cell(r,1).border=bdr(); ws_a.cell(r,1).alignment=lft()
ws_a.cell(r,2).value = ""; ws_a.cell(r,2).fill=fill(NAVY); ws_a.cell(r,2).border=bdr()
ws_a.cell(r,3).value = f"=SUM(C{blend_start}:C{r-1})"; ws_a.cell(r,3).font=fnt(bold=True,color=WHITE); ws_a.cell(r,3).fill=fill(NAVY); ws_a.cell(r,3).number_format=FMT_INT; ws_a.cell(r,3).border=bdr(); ws_a.cell(r,3).alignment=ctr()
ws_a.cell(r,4).value = f"=SUM(D{blend_start}:D{r-1})"; ws_a.cell(r,4).font=fnt(bold=True,color=WHITE); ws_a.cell(r,4).fill=fill(NAVY); ws_a.cell(r,4).number_format=FMT_USD; ws_a.cell(r,4).border=bdr(); ws_a.cell(r,4).alignment=ctr()
for ci in [5,6,7]:
    ws_a.cell(r,ci).fill=fill(NAVY); ws_a.cell(r,ci).border=bdr()

r += 2

# ── Section 6: JHBI Platform Migration Uplift ────────────────────────────────
hdr_row(ws_a, r, 7, "SECTION 6 — JHBI PLATFORM MIGRATION UPLIFT  (JHAKnow + ARCU → JHBI)", bg=COBALT); r+=1
ws_a.row_dimensions[r-1].height = 18

for ci, h in enumerate(["Parameter", "Value", "", "", "", "", "Notes"], 1):
    ws_a.cell(r,ci).value = h
    ws_a.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_a.cell(r,ci).fill  = fill(COBALT_D)
    ws_a.cell(r,ci).alignment = ctr()
    ws_a.cell(r,ci).border = bdr()
ws_a.row_dimensions[r].height = 18; r+=1

mig_start_r = r   # row of first mig_param (JHAKnow FI count)

mig_params = [
    ("JHAKnow Active FIs (Banks)",                    150,    FMT_INT, True,  "Current contracted bank FIs on JHAKnow platform"),
    ("ARCU Active FIs (CUs)",                          150,    FMT_INT, True,  "Current contracted CU FIs on ARCU platform"),
    ("Base Platform Price Assumption ($/FI/yr)",       10000,  FMT_USD, True,  "Assumed annual platform price for JHAKnow/ARCU — adjust as needed"),
    ("JHBI Platform Uplift %",                         0.10,   FMT_PCT, True,  "Nominal upcharge for migrating to JHBI; placeholder — not yet finalized"),
    ("Bank Upcharge ($/FI/yr)",                        None,   FMT_USD, False, "= Base Price × Uplift % — formula-driven"),
    ("CU Upcharge ($/FI/yr)",                          None,   FMT_USD, False, "= Bank Upcharge × (1 − CU Discount) — formula-driven"),
]

for i, (label, val, fmt, is_input, note) in enumerate(mig_params):
    bg = LT_BLUE if i % 2 == 0 else VLT_BLUE
    ws_a.row_dimensions[r].height = 18
    ws_a.cell(r,1).value = label; ws_a.cell(r,1).font = fnt(); ws_a.cell(r,1).fill = fill(bg); ws_a.cell(r,1).border = bdr(); ws_a.cell(r,1).alignment = lft()
    c = ws_a.cell(r,2)
    if val is None:
        if i == 4:   # Bank Upcharge = base_price × uplift_pct
            c.value = f"=B{mig_start_r+2}*B{mig_start_r+3}"
        elif i == 5: # CU Upcharge = bank_upcharge × (1 - CU discount at B8)
            c.value = f"=B{mig_start_r+4}*(1-$B$8)"
        c.font = fnt(color=FX_BLACK)
    else:
        c.value = val
        c.font  = fnt(color=IN_BLUE if is_input else FX_BLACK, bold=is_input)
    c.number_format = fmt
    c.fill   = fill(YELLOW if is_input else bg)
    c.border = bdr(); c.alignment = ctr()
    ws_a.cell(r,7).value = note; ws_a.cell(r,7).font = fnt(color="595959", size=9, italic=True); ws_a.cell(r,7).alignment = lft()
    r += 1

# Track row numbers needed by Revenue Model formulas
mig_bank_upcharge_row = mig_start_r + 4   # Bank Upcharge per FI/yr
mig_cu_upcharge_row   = mig_start_r + 5   # CU Upcharge per FI/yr

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — REVENUE MODEL
# ─────────────────────────────────────────────────────────────────────────────
ws_r = wb.create_sheet("Revenue Model")
ws_r.sheet_view.showGridLines = False
ws_r.freeze_panes = "C5"

ws_r.column_dimensions["A"].width = 32
ws_r.column_dimensions["B"].width = 14
for ci in range(3, 12):
    ws_r.column_dimensions[get_column_letter(ci)].width = 13

QUARTERS = [
    ("Q3 FY26", "Jan–Mar 2026"),
    ("Q4 FY26", "Apr–Jun 2026"),
    ("Q1 FY27", "Jul–Sep 2026"),
    ("Q2 FY27", "Oct–Dec 2026"),
    ("Q3 FY27", "Jan–Mar 2027"),
    ("Q4 FY27", "Apr–Jun 2027"),
    ("Q1 FY28", "Jul–Sep 2027"),
    ("Q2 FY28", "Oct–Dec 2027"),
]
N = len(QUARTERS)
DATA_COLS = list(range(3, 3+N))  # columns C through J

# Title
ws_r.row_dimensions[1].height = 30
ws_r.merge_cells(f"A1:{get_column_letter(3+N)}1")
c = ws_r["A1"]
c.value = "JHBI DS Apps — Quarterly ARR Build  |  Path to $1M"
c.font  = fnt(bold=True, color=WHITE, size=14)
c.fill  = fill(NAVY); c.alignment = lft()

ws_r.row_dimensions[2].height = 16
ws_r.merge_cells(f"A2:{get_column_letter(3+N)}2")
c = ws_r["A2"]
c.value = "   ARR = Cumulative Active FIs × Annual Subscription Price.  Churn applied quarterly (Annual Rate ÷ 4).  All inputs in blue; formulas in black."
c.font  = fnt(color="595959", size=9, italic=True); c.fill = fill(LT_GRAY); c.alignment = lft()

# Quarter / date headers
ws_r.row_dimensions[3].height = 20
ws_r.row_dimensions[4].height = 16

ws_r.cell(3,1).value = "Metric"; ws_r.cell(3,1).font=fnt(bold=True,color=WHITE); ws_r.cell(3,1).fill=fill(NAVY); ws_r.cell(3,1).border=bdr(); ws_r.cell(3,1).alignment=lft()
ws_r.cell(3,2).value = "App"; ws_r.cell(3,2).font=fnt(bold=True,color=WHITE); ws_r.cell(3,2).fill=fill(NAVY); ws_r.cell(3,2).border=bdr(); ws_r.cell(3,2).alignment=ctr()
ws_r.cell(4,1).value = ""; ws_r.cell(4,1).fill=fill(COBALT_D); ws_r.cell(4,1).border=bdr()
ws_r.cell(4,2).value = ""; ws_r.cell(4,2).fill=fill(COBALT_D); ws_r.cell(4,2).border=bdr()

for i,(q,cal) in enumerate(QUARTERS):
    col = DATA_COLS[i]
    ws_r.cell(3, col).value = q
    ws_r.cell(3, col).font  = fnt(bold=True, color=WHITE, size=10)
    ws_r.cell(3, col).fill  = fill(NAVY)
    ws_r.cell(3, col).border = bdr()
    ws_r.cell(3, col).alignment = ctr()

    ws_r.cell(4, col).value = cal
    ws_r.cell(4, col).font  = fnt(color=WHITE, size=8)
    ws_r.cell(4, col).fill  = fill(COBALT_D)
    ws_r.cell(4, col).border = bdr()
    ws_r.cell(4, col).alignment = ctr()

# ── App blocks ────────────────────────────────────────────────────────────────
# For each app: new FIs per quarter (input), active FIs (formula), ARR (formula)
# launch_q = 0-indexed quarter when app starts selling

apps_config = [
    # (name, blended_price_row_in_assumptions, launch_q, new_FIs_by_quarter, color)
    # new_FIs_by_quarter: list of 8 values (0 before launch)
    ("Zelle Memo Intelligence",  blend_start+0, 2,  [0, 0,  5,  4,  4,  4,  3,  3], "085CE5"),
    ("Churn Sentinel",           blend_start+1, 3,  [0, 0,  0, 15, 13, 12, 10,  8], "06185F"),
    ("CommercialSignal",         blend_start+2, 4,  [0, 0,  0,  0, 10,  9,  8,  7], "073FA8"),
    ("Gen. Wealth Deflection",   blend_start+3, 6,  [0, 0,  0,  0,  0,  0,  7,  6], "0D2E7A"),
    ("Anomaly Detection",        blend_start+4, 7,  [0, 0,  0,  0,  0,  0,  0,  7], "575A5D"),
]

# Track row numbers for ARR rows (used in totals)
arr_rows = []
new_fi_rows = []
active_fi_rows = []

r = 5

for app_i, (app_name, price_row, launch_q, new_fi_sched, app_color) in enumerate(apps_config):
    # Section header
    ws_r.row_dimensions[r].height = 20
    ws_r.merge_cells(f"A{r}:{get_column_letter(3+N)}{r}")
    c = ws_r.cell(r, 1)
    c.value = f"  {app_name}"
    c.font  = fnt(bold=True, color=WHITE, size=10)
    c.fill  = fill(app_color)
    c.border = bdr()
    c.alignment = lft()
    r += 1

    # Row 1: New FIs this quarter (INPUT — blue)
    ws_r.row_dimensions[r].height = 18
    ws_r.cell(r,1).value = "New FIs Signed"; ws_r.cell(r,1).font=fnt(); ws_r.cell(r,1).fill=fill(LT_BLUE); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
    ws_r.cell(r,2).value = "#"; ws_r.cell(r,2).font=fnt(); ws_r.cell(r,2).fill=fill(LT_BLUE); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
    new_fi_row = r
    new_fi_rows.append(r)
    for i, nf in enumerate(new_fi_sched):
        col = DATA_COLS[i]
        c = ws_r.cell(r, col)
        c.value = nf
        c.font  = fnt(color=IN_BLUE, bold=True) if nf > 0 else fnt(color="BFBFBF")
        c.fill  = fill(YELLOW) if nf > 0 else fill(LT_BLUE)
        c.border = bdr()
        c.alignment = ctr()
        c.number_format = FMT_INT
    r += 1

    # Row 2: Active FIs cumulative (formula: prev_active*(1-churn/4) + new)
    ws_r.row_dimensions[r].height = 18
    ws_r.cell(r,1).value = "Active FIs (Cumul., net churn)"; ws_r.cell(r,1).font=fnt(); ws_r.cell(r,1).fill=fill(VLT_BLUE); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
    ws_r.cell(r,2).value = "#"; ws_r.cell(r,2).font=fnt(); ws_r.cell(r,2).fill=fill(VLT_BLUE); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
    active_fi_row = r
    active_fi_rows.append(r)
    for i in range(N):
        col = DATA_COLS[i]
        c = ws_r.cell(r, col)
        if i == 0:
            c.value = f"=ROUND({get_column_letter(col)}{new_fi_row},0)"
        else:
            prev_col = get_column_letter(DATA_COLS[i-1])
            cur_col  = get_column_letter(col)
            c.value  = f"=ROUND({prev_col}{r}*(1-Assumptions!$B$10/4)+{cur_col}{new_fi_row},0)"
        c.font   = fnt()
        c.fill   = fill(VLT_BLUE)
        c.border = bdr()
        c.alignment = ctr()
        c.number_format = FMT_INT
    r += 1

    # Row 3: ARR (formula: active_FIs * blended_price)
    ws_r.row_dimensions[r].height = 18
    ws_r.cell(r,1).value = "Ending ARR"; ws_r.cell(r,1).font=fnt(bold=True); ws_r.cell(r,1).fill=fill(GREEN_BG); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
    ws_r.cell(r,2).value = "$"; ws_r.cell(r,2).font=fnt(bold=True); ws_r.cell(r,2).fill=fill(GREEN_BG); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
    arr_row = r
    arr_rows.append(r)
    for i in range(N):
        col = DATA_COLS[i]
        c = ws_r.cell(r, col)
        c.value  = f"=ROUND({get_column_letter(col)}{active_fi_row}*Assumptions!$B${price_row},0)"
        c.font   = fnt(bold=True)
        c.fill   = fill(GREEN_BG)
        c.border = bdr()
        c.alignment = ctr()
        c.number_format = FMT_USD
    r += 1
    r += 1  # blank separator

# ── MIGRATION UPLIFT BLOCK ────────────────────────────────────────────────────
MIG_TEAL = "1B7FA7"
ws_r.row_dimensions[r].height = 20
ws_r.merge_cells(f"A{r}:{get_column_letter(3+N)}{r}")
c = ws_r.cell(r,1)
c.value = "  JHBI PLATFORM MIGRATION UPLIFT — JHAKnow + ARCU → JHBI  (Nominal Upcharge)"
c.font  = fnt(bold=True, color=WHITE, size=10); c.fill = fill(MIG_TEAL); c.border = bdr(); c.alignment = lft()
r += 1

# Migration schedule: cumulative FIs migrated (hardcoded input — 4-quarter rollout Q1-Q4 FY27)
mig_bank_sched = [0, 0, 20, 75, 120, 150, 150, 150]   # banks (JHAKnow → JHBI)
mig_cu_sched   = [0, 0, 20, 75, 120, 150, 150, 150]   # CUs   (ARCU  → JHBI)

for row_label, sched, bg_color in [
    ("Cumul. Bank FIs Migrated (JHAKnow → JHBI)", mig_bank_sched, LT_BLUE),
    ("Cumul. CU FIs Migrated (ARCU → JHBI)",      mig_cu_sched,   VLT_BLUE),
]:
    ws_r.row_dimensions[r].height = 18
    ws_r.cell(r,1).value = row_label; ws_r.cell(r,1).font=fnt(); ws_r.cell(r,1).fill=fill(bg_color); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
    ws_r.cell(r,2).value = "#"; ws_r.cell(r,2).font=fnt(); ws_r.cell(r,2).fill=fill(bg_color); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
    if row_label.startswith("Cumul. Bank"):
        mig_bank_fi_row = r
    else:
        mig_cu_fi_row = r
    for i, nf in enumerate(sched):
        col = DATA_COLS[i]; c = ws_r.cell(r, col)
        c.value = nf
        c.font  = fnt(color=IN_BLUE, bold=True) if nf > 0 else fnt(color="BFBFBF")
        c.fill  = fill(YELLOW) if nf > 0 else fill(bg_color)
        c.border = bdr(); c.alignment = ctr(); c.number_format = FMT_INT
    r += 1

# Migration Upcharge ARR = banks × bank_upcharge/yr + CUs × cu_upcharge/yr (annual run rate)
ws_r.row_dimensions[r].height = 18
ws_r.cell(r,1).value = "Platform Migration Upcharge ARR"; ws_r.cell(r,1).font=fnt(bold=True); ws_r.cell(r,1).fill=fill(GREEN_BG); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
ws_r.cell(r,2).value = "$"; ws_r.cell(r,2).font=fnt(bold=True); ws_r.cell(r,2).fill=fill(GREEN_BG); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
mig_arr_row = r
for i in range(N):
    col = DATA_COLS[i]; cl = get_column_letter(col); c = ws_r.cell(r, col)
    c.value = f"=ROUND({cl}{mig_bank_fi_row}*Assumptions!$B${mig_bank_upcharge_row}+{cl}{mig_cu_fi_row}*Assumptions!$B${mig_cu_upcharge_row},0)"
    c.font = fnt(bold=True); c.fill = fill(GREEN_BG); c.border = bdr(); c.alignment = ctr(); c.number_format = FMT_USD
r += 2

# ── TOTAL ARR block ───────────────────────────────────────────────────────────
ws_r.row_dimensions[r].height = 22
ws_r.merge_cells(f"A{r}:{get_column_letter(3+N)}{r}")
c = ws_r.cell(r,1)
c.value = "  TOTAL — COMBINED JHBI ARR  (DS Apps + Platform Migration Uplift)"
c.font  = fnt(bold=True, color=WHITE, size=11)
c.fill  = fill(NAVY); c.border = bdr(); c.alignment = lft()
r += 1

# Total active FIs
ws_r.row_dimensions[r].height = 18
ws_r.cell(r,1).value = "Total Active FI App Subscriptions"; ws_r.cell(r,1).font=fnt(bold=True); ws_r.cell(r,1).fill=fill(LT_BLUE); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
ws_r.cell(r,2).value = "#"; ws_r.cell(r,2).font=fnt(bold=True); ws_r.cell(r,2).fill=fill(LT_BLUE); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
total_fi_row = r
for i in range(N):
    col = DATA_COLS[i]
    sum_formula = "+".join([f"{get_column_letter(col)}{ar}" for ar in active_fi_rows])
    c = ws_r.cell(r, col)
    c.value = f"={sum_formula}"
    c.font  = fnt(bold=True); c.fill = fill(LT_BLUE); c.border = bdr(); c.alignment = ctr(); c.number_format = FMT_INT
r += 1

# Total ARR
ws_r.row_dimensions[r].height = 22
ws_r.cell(r,1).value = "TOTAL ARR"; ws_r.cell(r,1).font=fnt(bold=True,color=WHITE,size=11); ws_r.cell(r,1).fill=fill(COBALT); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
ws_r.cell(r,2).value = "$"; ws_r.cell(r,2).font=fnt(bold=True,color=WHITE); ws_r.cell(r,2).fill=fill(COBALT); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
total_arr_row = r
for i in range(N):
    col = DATA_COLS[i]
    sum_formula = "+".join([f"{get_column_letter(col)}{ar}" for ar in arr_rows]) + f"+{get_column_letter(col)}{mig_arr_row}"
    c = ws_r.cell(r, col)
    c.value  = f"={sum_formula}"
    c.font   = fnt(bold=True, color=WHITE, size=11)
    c.fill   = fill(COBALT)
    c.border = bdr()
    c.alignment = ctr()
    c.number_format = FMT_USD
r += 1

# QoQ Growth
ws_r.row_dimensions[r].height = 18
ws_r.cell(r,1).value = "QoQ ARR Growth"; ws_r.cell(r,1).font=fnt(); ws_r.cell(r,1).fill=fill(LT_GRAY); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
ws_r.cell(r,2).value = "%"; ws_r.cell(r,2).font=fnt(); ws_r.cell(r,2).fill=fill(LT_GRAY); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
for i in range(N):
    col = DATA_COLS[i]
    c = ws_r.cell(r, col)
    if i == 0:
        c.value = '"-"'
    else:
        prev_col = get_column_letter(DATA_COLS[i-1])
        c.value  = f'=IF({prev_col}{total_arr_row}=0,"-",({get_column_letter(col)}{total_arr_row}-{prev_col}{total_arr_row})/{prev_col}{total_arr_row})'
    c.font   = fnt(); c.fill = fill(LT_GRAY); c.border = bdr(); c.alignment = ctr()
    c.number_format = FMT_PCT

r += 1

# $1M milestone marker
ws_r.row_dimensions[r].height = 18
ws_r.cell(r,1).value = "$1M ARR Milestone"; ws_r.cell(r,1).font=fnt(bold=True); ws_r.cell(r,1).fill=fill(AMBER); ws_r.cell(r,1).border=bdr(); ws_r.cell(r,1).alignment=lft()
ws_r.cell(r,2).value = "✓"; ws_r.cell(r,2).font=fnt(bold=True); ws_r.cell(r,2).fill=fill(AMBER); ws_r.cell(r,2).border=bdr(); ws_r.cell(r,2).alignment=ctr()
for i in range(N):
    col = DATA_COLS[i]
    c = ws_r.cell(r, col)
    c.value  = f'=IF({get_column_letter(col)}{total_arr_row}>=1000000,"✓ $1M","—")'
    c.font   = fnt(bold=True)
    c.fill   = fill(AMBER)
    c.border = bdr()
    c.alignment = ctr()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — SUMMARY: PATH TO $1M
# ─────────────────────────────────────────────────────────────────────────────
ws_s = wb.create_sheet("Summary — Path to $1M")
ws_s.sheet_view.showGridLines = False
ws_s.column_dimensions["A"].width = 18
ws_s.column_dimensions["B"].width = 18
ws_s.column_dimensions["C"].width = 18
ws_s.column_dimensions["D"].width = 18
ws_s.column_dimensions["E"].width = 18
ws_s.column_dimensions["F"].width = 18
ws_s.column_dimensions["G"].width = 20
ws_s.column_dimensions["H"].width = 22
ws_s.column_dimensions["I"].width = 26

# Title
ws_s.row_dimensions[1].height = 30
ws_s.merge_cells("A1:H1")
c = ws_s["A1"]
c.value = "JHBI DS Apps  —  Path to $1M ARR"
c.font  = fnt(bold=True, color=WHITE, size=14); c.fill = fill(NAVY); c.alignment = lft()

ws_s.row_dimensions[2].height = 16
ws_s.merge_cells("A2:H2")
c = ws_s["A2"]
c.value = "   Combined JHBI ARR: DS Apps + Platform Migration Uplift (JHAKnow + ARCU → JHBI).  Milestone: Q4 FY27 combined.  All values from Revenue Model."
c.font  = fnt(color="595959", size=9, italic=True); c.fill = fill(LT_GRAY); c.alignment = lft()

r = 4

# Quarter summary table
hdr_row(ws_s, r, 8, "QUARTERLY ARR SUMMARY", bg=COBALT); r+=1
ws_s.row_dimensions[r-1].height = 18

for ci, h in enumerate(["Quarter","Calendar","Zelle Memo ARR","Churn Sentinel ARR","CommercialSignal ARR","Gen. Wealth ARR","Anomaly ARR","Platform Migration ARR","Total JHBI ARR"], 1):
    ws_s.cell(r,ci).value = h
    ws_s.cell(r,ci).font  = fnt(bold=True, color=WHITE, size=9)
    ws_s.cell(r,ci).fill  = fill(COBALT_D)
    ws_s.cell(r,ci).alignment = ctr()
    ws_s.cell(r,ci).border = bdr()
ws_s.row_dimensions[r].height = 18; r+=1

summary_start = r
rev_arr_rows = arr_rows  # from Revenue Model build
rev_total_row = total_arr_row

for i, (q, cal) in enumerate(QUARTERS):
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_s.row_dimensions[r].height = 18
    col_letter = get_column_letter(DATA_COLS[i])
    ws_s.cell(r,1).value = q;   ws_s.cell(r,1).font=fnt(bold=(i==5)); ws_s.cell(r,1).fill=fill(AMBER if i==5 else bg); ws_s.cell(r,1).border=bdr(); ws_s.cell(r,1).alignment=ctr()
    ws_s.cell(r,2).value = cal; ws_s.cell(r,2).font=fnt(size=9); ws_s.cell(r,2).fill=fill(AMBER if i==5 else bg); ws_s.cell(r,2).border=bdr(); ws_s.cell(r,2).alignment=ctr()

    for ci_offset, arr_row_num in enumerate(rev_arr_rows):
        c = ws_s.cell(r, 3+ci_offset)
        c.value  = f"='Revenue Model'!{col_letter}{arr_row_num}"
        c.font   = fnt(color=XS_GREEN, bold=(i==5))
        c.fill   = fill(AMBER if i==5 else bg)
        c.border = bdr()
        c.alignment = ctr()
        c.number_format = FMT_USD

    # Platform Migration ARR (column 8)
    c = ws_s.cell(r, 8)
    c.value  = f"='Revenue Model'!{col_letter}{mig_arr_row}"
    c.font   = fnt(color=XS_GREEN, bold=(i==5))
    c.fill   = fill(AMBER if i==5 else bg)
    c.border = bdr(); c.alignment = ctr(); c.number_format = FMT_USD

    # Total JHBI ARR (column 9) — references Revenue Model total which includes migration
    c = ws_s.cell(r, 9)
    c.value  = f"='Revenue Model'!{col_letter}{rev_total_row}"
    c.font   = fnt(bold=True, color=(WHITE if i==5 else FX_BLACK))
    c.fill   = fill(COBALT if i==5 else bg)
    c.border = bdr(); c.alignment = ctr(); c.number_format = FMT_USD

    if i == 5:
        ws_s.row_dimensions[r].height = 22
    r += 1

# Grand total / milestone note
ws_s.row_dimensions[r].height = 18
for ci in range(1, 10):
    c = ws_s.cell(r, ci)
    c.fill = fill(NAVY); c.border = bdr()
ws_s.cell(r,1).value = "★ Combined JHBI ARR (DS Apps + Migration) crosses $1M in Q4 FY27 (Apr–Jun 2027)  |  DS Apps alone: $1M in Q1 FY28  |  Migration: 10% uplift placeholder on 300 JHAKnow/ARCU FIs"; ws_s.cell(r,1).font=fnt(bold=True,color=WHITE); ws_s.cell(r,1).alignment=lft()
ws_s.merge_cells(f"A{r}:I{r}")
r += 2

# Key assumptions recap
hdr_row(ws_s, r, 9, "KEY ASSUMPTIONS DRIVING THE MODEL", bg=COBALT); r+=1
ws_s.row_dimensions[r-1].height = 18

recap = [
    ("Total JH Addressable FIs",        "1,660",       "Contracted core banking FIs (44% CU / 56% Bank)"),
    ("CU Share",                        "44%",         "Symitar / credit union base"),
    ("CU Discount",                     "12%",         "Cooperative budget accommodation"),
    ("Zelle Penetration",               "15.7%",       "260 of ~1,660 JH core FIs currently Zelle-enabled"),
    ("Annual DS App Churn",             "8%",          "FI churn on DS app subscriptions (2% quarterly)"),
    ("Migration FIs (JHAKnow + ARCU)", "300",          "150 bank (JHAKnow) + 150 CU (ARCU); nominal uplift placeholder"),
    ("Platform Uplift %",               "10%",         "Placeholder — not finalized; adjust in Section 6 of Assumptions"),
    ("★ Combined $1M Milestone",        "Q4 FY27",     "DS Apps ($910K) + Platform Migration ($282K) = $1.19M ARR"),
    ("DS Apps Alone $1M Milestone",     "Q1 FY28",     "$1.25M ARR from 4 apps, 102 active FI subscriptions, 6.1% penetration"),
]
for i,(label,val,note) in enumerate(recap):
    bg = LT_BLUE if i%2==0 else VLT_BLUE
    ws_s.row_dimensions[r].height = 18
    ws_s.cell(r,1).value=label; ws_s.cell(r,1).font=fnt(); ws_s.cell(r,1).fill=fill(bg); ws_s.cell(r,1).border=bdr(); ws_s.cell(r,1).alignment=lft()
    ws_s.merge_cells(f"A{r}:D{r}")
    ws_s.cell(r,5).value=val;   ws_s.cell(r,5).font=fnt(bold=True); ws_s.cell(r,5).fill=fill(GREEN_BG); ws_s.cell(r,5).border=bdr(); ws_s.cell(r,5).alignment=ctr()
    ws_s.cell(r,6).value=note;  ws_s.cell(r,6).font=fnt(color="595959",size=9,italic=True); ws_s.cell(r,6).fill=fill(bg); ws_s.cell(r,6).border=bdr(); ws_s.cell(r,6).alignment=lft()
    ws_s.merge_cells(f"F{r}:I{r}")
    r += 1

# Sheet order
wb.move_sheet("Summary — Path to $1M", offset=-2)

out = "/sessions/laughing-keen-fermat/mnt/outputs/JHBI_Revenue_Model.xlsx"
wb.save(out)
print(f"Saved: {out}")
